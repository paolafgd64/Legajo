"""Vistas del modulo de administracion.

Se concentra aqui todo lo relacionado con panel admin, carga masiva y
seguimiento de reportes de usuarios.
"""

import json
from datetime import timedelta
from json import JSONDecodeError
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.db import DatabaseError
from django.db.models import Count, Q
from django.http import JsonResponse
from django.shortcuts import redirect, render
from django.utils import timezone
from django.views.decorators.cache import never_cache
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.decorators.http import require_http_methods

from ..gestion_libros.models import Libro
from ..intercambios.models import Intercambio
from ..services import import_users_from_payload, serialize_book
from ..validators import ControlledError
from ..views.helpers import _read_json_body
from .helpers import (
    _admin_required_response,
    _build_monthly_cumulative_series,
    _get_admin_dashboard_context,
    _serialize_admin_user,
)
from .models import ConfiguracionContacto, NotificacionUsuario, ReporteUsuario


User = get_user_model()


def _calcular_fin_suspension(data):
    try:
        valor = int(data.get('duracionValor') or 0)
    except (TypeError, ValueError):
        return None, 'Indica una duracion de suspension valida.'

    unidad = str(data.get('duracionUnidad') or 'dias').strip().lower()
    if valor <= 0:
        return None, 'La duracion de la suspension debe ser mayor a cero.'
    if unidad not in {'horas', 'dias'}:
        return None, 'La unidad de suspension no es valida.'

    delta = timedelta(hours=valor) if unidad == 'horas' else timedelta(days=valor)
    if delta > timedelta(days=365):
        return None, 'La suspension no puede superar 365 dias.'

    return timezone.now() + delta, ''


def _wants_json_response(request):
    accept = request.headers.get('Accept', '')
    requested_with = request.headers.get('X-Requested-With', '')
    return 'application/json' in accept or requested_with == 'XMLHttpRequest'


def _parse_excel_payload(archivo):
    try:
        from openpyxl import load_workbook

        archivo.seek(0)
        workbook = load_workbook(archivo, read_only=True, data_only=True)
    except Exception as exc:
        raise ControlledError('No se pudo leer el archivo Excel. Verifica que sea un .xlsx valido.', status_code=400) from exc

    hoja = workbook.active
    filas = list(hoja.iter_rows(values_only=True))
    if not filas:
        raise ControlledError('El archivo Excel esta vacio.', status_code=400)

    encabezados = [str(valor).strip() if valor is not None else '' for valor in filas[0]]
    if not any(encabezados):
        raise ControlledError('La primera fila del archivo Excel debe contener encabezados.', status_code=400)

    payload = []
    for fila in filas[1:]:
        if not fila or not any(valor not in (None, '') for valor in fila):
            continue

        item = {}
        for indice, encabezado in enumerate(encabezados):
            if not encabezado:
                continue
            valor = fila[indice] if indice < len(fila) else None
            item[encabezado] = '' if valor is None else valor
        payload.append(item)

    if not payload:
        raise ControlledError('El archivo Excel no contiene filas de datos para importar.', status_code=400)

    return payload


def _load_users_payload_from_file(archivo):
    nombre = (getattr(archivo, 'name', '') or '').lower()
    if nombre.endswith('.xlsx'):
        return _parse_excel_payload(archivo)

    try:
        archivo.seek(0)
        return json.loads(archivo.read().decode('utf-8-sig'))
    except UnicodeDecodeError as exc:
        raise ControlledError('El archivo debe estar codificado en UTF-8 o ser un Excel .xlsx valido.', status_code=400) from exc
    except JSONDecodeError as exc:
        raise ControlledError(f'El archivo no contiene JSON valido: {exc}', status_code=400) from exc


@login_required(login_url='login')
@never_cache
def dashboard_admin(request):
    if request.user.rol != User.Rol.ADMIN:
        return redirect('dashboard_usuario')

    context = _get_admin_dashboard_context()
    context['admin_name'] = request.user.nombre1 or 'Administrador'
    return render(request, 'administracion/dashboard_admin.html', context)


@ensure_csrf_cookie
@login_required(login_url='login')
@never_cache
@require_http_methods(["GET", "POST"])
def usuarios_admin(request):
    admin_error = _admin_required_response(request)
    if admin_error:
        return redirect('dashboard_usuario') if request.user.is_authenticated else redirect('login')

    context = {
        'admin_name': request.user.nombre1 or 'Administrador',
        'total_usuarios_cargados': User.objects.filter(
            rol=User.Rol.USUARIO,
            activo=True,
            is_active=True,
        ).count(),
    }

    mensaje_exito = request.session.pop('mensaje_exito_carga_usuarios', None)
    if mensaje_exito:
        context['mensaje_exito_carga'] = mensaje_exito

    if request.method == 'POST':
        archivo = request.FILES.get('archivo_usuarios')
        actualizar = request.POST.get('actualizar_existentes') == 'on'

        if not archivo:
            message = 'Debes seleccionar un archivo JSON o Excel para importar.'
            if _wants_json_response(request):
                return JsonResponse({'message': message}, status=400)
            context['mensaje_error_carga'] = message
            return render(request, 'administracion/usuarios_admin.html', context, status=400)

        try:
            # La vista solo parsea el archivo; la regla de negocio de creacion
            # y actualizacion vive en services/users.py.
            payload = _load_users_payload_from_file(archivo)
            resultado = import_users_from_payload(payload, actualizar=actualizar)
        except ControlledError as exc:
            if _wants_json_response(request):
                return JsonResponse({'message': exc.message}, status=exc.status_code)
            context['mensaje_error_carga'] = exc.message
            return render(request, 'administracion/usuarios_admin.html', context, status=400)

        success_message = (
            f"Importacion completada. Creados: {resultado['creados']}, "
            f"actualizados: {resultado['actualizados']}, omitidos: {resultado['omitidos']}."
        )
        if _wants_json_response(request):
            return JsonResponse({
                'message': success_message,
                'resultado': resultado,
            })

        request.session['mensaje_exito_carga_usuarios'] = success_message
        return redirect('usuarios_admin')

    return render(request, 'administracion/usuarios_admin.html', context)


@ensure_csrf_cookie
@login_required(login_url='login')
@never_cache
@require_http_methods(["GET", "POST"])
def carga_masiva_usuarios(request):
    admin_error = _admin_required_response(request)
    if admin_error:
        return redirect('dashboard_usuario') if request.user.is_authenticated else redirect('login')

    context = {
        'admin_name': request.user.nombre1 or 'Administrador',
    }

    if request.method == 'POST':
        archivo = request.FILES.get('archivo_usuarios')
        actualizar = request.POST.get('actualizar_existentes') == 'on'
        context['actualizar_existentes'] = actualizar

        if not archivo:
            context['error_message'] = 'Debes seleccionar un archivo JSON o Excel para importar.'
            return render(request, 'administracion/carga_masiva_usuarios.html', context, status=400)

        try:
            payload = _load_users_payload_from_file(archivo)
            resultado = import_users_from_payload(payload, actualizar=actualizar)
        except ControlledError as exc:
            context['error_message'] = exc.message
            return render(request, 'administracion/carga_masiva_usuarios.html', context, status=400)

        context['success_message'] = (
            f"Importacion completada. Creados: {resultado['creados']}, "
            f"actualizados: {resultado['actualizados']}, omitidos: {resultado['omitidos']}."
        )

    return render(request, 'administracion/carga_masiva_usuarios.html', context)


@login_required(login_url='login')
@never_cache
def inventario_admi_redirect(request):
    if request.user.rol != User.Rol.ADMIN:
        return redirect('dashboard_usuario')
    return redirect('dashboard_admin')


@login_required(login_url='login')
@never_cache
def perfil_admin(request):
    user = request.user
    return render(request, 'administracion/perfil_admin.html', {
        'perfil_usuario': user,
        'nombre_completo': ' '.join(filter(None, [user.nombre1, user.nombre2, user.apellido1, user.apellido2])),
        'contacto_sitio': ConfiguracionContacto.obtener(),
    })


@login_required(login_url='login')
@require_http_methods(["GET", "PUT"])
def api_admin_contact_config(request):
    admin_error = _admin_required_response(request)
    if admin_error:
        return admin_error

    contacto = ConfiguracionContacto.obtener()
    if request.method == 'PUT':
        data = _read_json_body(request)
        if data is None:
            return JsonResponse({'message': 'El cuerpo de la solicitud no es JSON valido.'}, status=400)

        telefono = str(data.get('telefono') or '').strip()
        whatsapp = str(data.get('whatsapp') or '').strip()
        correo = str(data.get('correo') or '').strip().lower()

        if not telefono or not whatsapp or not correo:
            return JsonResponse({'message': 'Telefono, WhatsApp y correo son obligatorios.'}, status=400)
        if len(telefono) > 30 or len(whatsapp) > 30:
            return JsonResponse({'message': 'Telefono y WhatsApp no deben superar 30 caracteres.'}, status=400)
        if not all(char.isdigit() or char in '+ ()-' for char in telefono + whatsapp):
            return JsonResponse({'message': 'Telefono y WhatsApp solo deben contener numeros, espacios, +, guiones o parentesis.'}, status=400)
        try:
            validate_email(correo)
        except ValidationError:
            return JsonResponse({'message': 'Ingresa un correo electronico valido.'}, status=400)

        contacto.telefono = telefono
        contacto.whatsapp = whatsapp
        contacto.correo = correo
        contacto.save(update_fields=['telefono', 'whatsapp', 'correo', 'fecha_actualizacion'])

    return JsonResponse({
        'telefono': contacto.telefono,
        'whatsapp': contacto.whatsapp,
        'correo': contacto.correo,
    })


@ensure_csrf_cookie
@login_required(login_url='login')
@never_cache
def novedades_usuarios(request):
    admin_error = _admin_required_response(request)
    if admin_error:
        return redirect('dashboard_usuario')
    return render(request, 'administracion/novedades_usuarios.html', {
        'admin_name': request.user.nombre1 or 'Administrador',
    })


@login_required(login_url='login')
@require_http_methods(["GET"])
def api_admin_reported_users(request):
    admin_error = _admin_required_response(request)
    if admin_error:
        return admin_error

    data = _build_monthly_cumulative_series(
        ReporteUsuario.objects.filter(activo=True),
        'fecha_reporte',
    )
    return JsonResponse(data, safe=False)


@login_required(login_url='login')
@require_http_methods(["GET"])
def api_admin_completed_exchanges(request):
    admin_error = _admin_required_response(request)
    if admin_error:
        return admin_error

    data = _build_monthly_cumulative_series(
        Intercambio.objects.filter(
            activo=True,
            estado=Intercambio.Estado.COMPLETADO,
            fecha_completado__isnull=False,
        ),
        'fecha_completado',
    )
    return JsonResponse(data, safe=False)


@login_required(login_url='login')
@require_http_methods(["GET"])
def api_admin_users(request):
    admin_error = _admin_required_response(request)
    if admin_error:
        return admin_error

    usuarios = User.objects.all().order_by('-id')

    nombre = (request.GET.get('nombre') or '').strip()
    correo = (request.GET.get('correo') or '').strip()
    ciudad = (request.GET.get('ciudad') or '').strip()
    estado = (request.GET.get('estado') or '').strip()

    if nombre:
        usuarios = usuarios.filter(
            Q(nombre1__icontains=nombre) |
            Q(nombre2__icontains=nombre) |
            Q(apellido1__icontains=nombre) |
            Q(apellido2__icontains=nombre)
        )
    if correo:
        usuarios = usuarios.filter(email__icontains=correo)
    if ciudad:
        usuarios = usuarios.filter(ciudad__icontains=ciudad)
    if estado == 'activo':
        usuarios = usuarios.filter(activo=True, is_active=True)
    elif estado == 'inactivo':
        usuarios = usuarios.filter(Q(activo=False) | Q(is_active=False))

    return JsonResponse([_serialize_admin_user(usuario) for usuario in usuarios], safe=False)


@login_required(login_url='login')
@require_http_methods(["GET"])
def api_admin_user_inventory(request, user_id):
    admin_error = _admin_required_response(request)
    if admin_error:
        return admin_error

    try:
        usuario = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'message': 'Usuario no encontrado.'}, status=404)

    libros = (
        Libro.objects.filter(usuario_propietario=usuario, activo=True)
        .prefetch_related('autores', 'generos')
        .order_by('-id')
    )

    return JsonResponse({
        'usuario': _serialize_admin_user(usuario),
        'libros': [serialize_book(libro) for libro in libros],
    })


@login_required(login_url='login')
@require_http_methods(["PATCH"])
def api_admin_update_user_status(request, user_id):
    admin_error = _admin_required_response(request)
    if admin_error:
        return admin_error

    data = _read_json_body(request)
    if data is None:
        return JsonResponse({'message': 'El cuerpo de la solicitud no es JSON valido.'}, status=400)

    activo = data.get('activo')
    if not isinstance(activo, bool):
        return JsonResponse({'message': 'El estado enviado no es valido.'}, status=400)

    try:
        usuario = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'message': 'Usuario no encontrado.'}, status=404)

    if usuario.id == request.user.id and not activo:
        return JsonResponse({'message': 'No puedes suspender o bloquear tu propia cuenta de administrador.'}, status=400)

    if activo:
        usuario.activo = True
        usuario.is_active = True
        usuario.motivo_desactivacion = ''
        usuario.suspension_hasta = None
    else:
        motivo = str(data.get('motivo') or '').strip()
        if len(motivo) < 10:
            return JsonResponse({'message': 'Debes escribir un motivo de al menos 10 caracteres.'}, status=400)

        accion = str(data.get('accion') or 'bloquear').strip().lower()
        if accion not in {'suspender', 'bloquear'}:
            return JsonResponse({'message': 'La accion enviada no es valida.'}, status=400)

        usuario.activo = False
        usuario.is_active = False
        usuario.motivo_desactivacion = motivo
        usuario.suspension_hasta = None

        if accion == 'suspender':
            suspension_hasta, error = _calcular_fin_suspension(data)
            if error:
                return JsonResponse({'message': error}, status=400)
            usuario.suspension_hasta = suspension_hasta

    try:
        usuario.save(update_fields=['activo', 'is_active', 'motivo_desactivacion', 'suspension_hasta'])
    except DatabaseError:
        return JsonResponse({'message': 'No fue posible actualizar el estado del usuario.'}, status=500)

    if not activo and usuario.suspension_hasta:
        mensaje = f'Usuario suspendido hasta {timezone.localtime(usuario.suspension_hasta).strftime("%Y-%m-%d %H:%M")}.'
    elif not activo:
        mensaje = 'Usuario bloqueado correctamente.'
    else:
        mensaje = 'Usuario activado correctamente.'

    return JsonResponse({
        'message': mensaje,
        'usuario': _serialize_admin_user(usuario),
    })


def _serialize_report(report):
    reportante = report.usuario_reportante
    reportado = report.usuario_reportado
    libro = report.libro_reportado
    return {
        'id': report.id,
        'motivo': report.motivo,
        'descripcion': report.descripcion,
        'estado': report.estado,
        'fechaReporte': report.fecha_reporte.strftime('%Y-%m-%d %H:%M') if report.fecha_reporte else '',
        'libroReportado': libro.titulo if libro else 'No especificado',
        'libroReportadoId': report.libro_reportado_id,
        'usuarioReportante': str(reportante) if reportante else 'Usuario desconocido',
        'usuarioReportanteId': report.usuario_reportante_id,
        'usuarioReportado': str(reportado) if reportado else 'Usuario desconocido',
        'usuarioReportadoId': report.usuario_reportado_id,
    }


@login_required(login_url='login')
@require_http_methods(["GET"])
def api_admin_user_reports(request):
    admin_error = _admin_required_response(request)
    if admin_error:
        return admin_error

    estado = (request.GET.get('estado') or '').strip().lower()
    busqueda = (request.GET.get('q') or '').strip()
    motivo = (request.GET.get('motivo') or '').strip()

    reportes = (
        ReporteUsuario.objects.filter(activo=True)
        .select_related('usuario_reportante', 'usuario_reportado', 'libro_reportado')
        .order_by('-fecha_reporte')
    )

    estados_validos = {value for value, _ in ReporteUsuario.Estado.choices}
    if estado in estados_validos:
        reportes = reportes.filter(estado=estado)

    if motivo:
        reportes = reportes.filter(motivo__icontains=motivo)

    if busqueda:
        reportes = reportes.filter(
            Q(descripcion__icontains=busqueda) |
            Q(motivo__icontains=busqueda) |
            Q(usuario_reportante__nombre1__icontains=busqueda) |
            Q(usuario_reportante__apellido1__icontains=busqueda) |
            Q(usuario_reportante__email__icontains=busqueda) |
            Q(usuario_reportado__nombre1__icontains=busqueda) |
            Q(usuario_reportado__apellido1__icontains=busqueda) |
            Q(usuario_reportado__email__icontains=busqueda) |
            Q(libro_reportado__titulo__icontains=busqueda)
        )

    return JsonResponse([_serialize_report(reporte) for reporte in reportes], safe=False)


@login_required(login_url='login')
@require_http_methods(["PATCH"])
def api_admin_update_user_report(request, report_id):
    admin_error = _admin_required_response(request)
    if admin_error:
        return admin_error

    data = _read_json_body(request)
    if data is None:
        return JsonResponse({'message': 'El cuerpo de la solicitud no es JSON valido.'}, status=400)

    nuevo_estado = str(data.get('estado') or '').strip().lower()
    estados_editables = {
        ReporteUsuario.Estado.REVISADO,
        ReporteUsuario.Estado.DESCARTADO,
    }
    if nuevo_estado not in estados_editables:
        return JsonResponse({'message': 'El estado enviado no es valido.'}, status=400)

    mensaje_respuesta = str(data.get('mensaje') or '').strip()
    if 'mensaje' in data and len(mensaje_respuesta) < 10:
        return JsonResponse({'message': 'Debes escribir una respuesta de al menos 10 caracteres.'}, status=400)

    try:
        reporte = ReporteUsuario.objects.get(id=report_id, activo=True)
    except ReporteUsuario.DoesNotExist:
        return JsonResponse({'message': 'Reporte no encontrado.'}, status=404)

    reporte.estado = nuevo_estado
    reporte.save(update_fields=['estado'])

    if reporte.usuario_reportante_id:
        # Si la tabla de notificaciones existe, avisamos al usuario que creo
        # el reporte para cerrar el ciclo de seguimiento.
        nombre_reportado = str(reporte.usuario_reportado) if reporte.usuario_reportado else 'el usuario reportado'
        if nuevo_estado == ReporteUsuario.Estado.REVISADO:
            mensaje_default = (
                f'Tu reporte sobre {nombre_reportado} ya fue revisado por el equipo administrador '
                'y se tomaron las medidas correspondientes.'
            )
            prefijo_respuesta = 'Tu reporte fue revisado y confirmado por el equipo administrador.'
        else:
            mensaje_default = (
                f'Tu reporte sobre {nombre_reportado} fue descartado porque no se encontraron '
                'motivos justificables para tomar medidas.'
            )
            prefijo_respuesta = 'Tu reporte fue descartado por el equipo administrador.'
        mensaje = f'{prefijo_respuesta} Respuesta: {mensaje_respuesta}' if mensaje_respuesta else mensaje_default

        try:
            NotificacionUsuario.objects.create(
                usuario_id=reporte.usuario_reportante_id,
                mensaje=mensaje,
                reporte_relacionado=reporte,
            )
        except DatabaseError:
            pass

    return JsonResponse({
        'message': 'Estado del reporte actualizado correctamente.',
        'reporte': _serialize_report(reporte),
    })
